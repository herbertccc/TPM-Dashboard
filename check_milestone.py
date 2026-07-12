#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('milestone/里程碑节点.xlsx', read_only=True, data_only=True)
print('Sheets:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(f'\n{sheet_name} ({len(rows)-1} rows):')
    for row in rows[:5]:
        print(' ', row)
wb.close()
