#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# TPM Dashboard Build Script
# Reads data from DingTalk API, computes metrics, injects into index.html
import csv
import json
import os
import re
import ssl
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    openpyxl = None

DATA_DIR = 'data'
DINGTALK_APP_KEY = os.environ.get('DINGTALK_APP_KEY', '')
DINGTALK_APP_SECRET = os.environ.get('DINGTALK_APP_SECRET', '')
DINGTALK_NODE_ID = 'P0MALyR8kNnB9yZliY70z99KJ3bzYmDO'
DINGTALK_OPERATOR_ID = os.environ.get('DINGTALK_OPERATOR_ID', 'PRmQRCFpuKchRpqCviSdoiiwiEiE')
TREND_DAYS = 30

SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

CLOSED = '\u5df2\u5173\u95ed'
RESOLVED = '\u5df2\u89e3\u51b3'
REOPEN = '\u91cd\u65b0\u6253\u5f00'
PENDING = '\u5f85\u89e3\u51b3'
CLOSED_SET = (CLOSED, RESOLVED)
OPEN_SET = (PENDING,)  # OPEN DI: only \u5f85\u89e3\u51b3
HIGH_RISK = '\u9ad8\u98ce\u9669\u95ee\u9898'
OVER90 = '\u8d8590\u5929'
UNKNOWN = '\u672a\u77e5'
IN_PROGRESS = '\u8fdb\u884c\u4e2d'
GENERAL = '\u901a\u7528'


def dingtalk_get_token():
    url = ('https://oapi.dingtalk.com/gettoken?appkey='
           + DINGTALK_APP_KEY + '&appsecret=' + DINGTALK_APP_SECRET)
    req = urllib.request.Request(url)
    with urllib.request.urlopen(req, timeout=15, context=SSL_CTX) as resp:
        data = json.loads(resp.read().decode('utf-8'))
        if data.get('errcode') == 0:
            print('token ok')
            return data['access_token']
        raise Exception('token failed: ' + str(data))


def _api_get(url, token, timeout=15):
    headers = {'x-acs-dingtalk-access-token': token}
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout, context=SSL_CTX) as resp:
        return json.loads(resp.read().decode('utf-8'))


def _api_post(url, token, body, timeout=30):
    headers = {
        'x-acs-dingtalk-access-token': token,
        'Content-Type': 'application/json'
    }
    data = json.dumps(body).encode('utf-8') if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method='POST')
    with urllib.request.urlopen(req, timeout=timeout, context=SSL_CTX) as resp:
        return json.loads(resp.read().decode('utf-8'))


def dingtalk_get_node_info(access_token):
    """Get knowledge base node info using multiple API strategies."""
    oid = DINGTALK_OPERATOR_ID
    # Strategy 1: Wiki API - GET /v2.0/wiki/nodes/{nodeId}
    if oid:
        try:
            url = ('https://api.dingtalk.com/v2.0/wiki/nodes/'
                   + DINGTALK_NODE_ID + '?operatorId=' + oid)
            data = _api_get(url, access_token)
            node = data.get('node', data)
            print('wiki node ok, type=' + node.get('type', '?')
                  + ', cat=' + node.get('category', '?'))
            return node
        except Exception as e:
            print('wiki nodes failed: ' + str(e))
    # Strategy 2: Doc dentry API - GET /v2.0/doc/dentries/{uuid}/queryDentryId
    if oid:
        try:
            url = ('https://api.dingtalk.com/v2.0/doc/dentries/'
                   + DINGTALK_NODE_ID + '/queryDentryId?operatorId=' + oid)
            data = _api_get(url, access_token)
            print('dentry ok: spaceId=' + data.get('spaceId', '?')
                  + ', dentryId=' + data.get('dentryId', '?'))
            return data
        except Exception as e:
            print('dentry query failed: ' + str(e))
    # Strategy 3: Drive spaces - GET /v1.0/drive/spaces/{spaceId}/files/{fileId}
    try:
        url = ('https://api.dingtalk.com/v1.0/drive/spaces/'
               + DINGTALK_NODE_ID + '/files/' + DINGTALK_NODE_ID)
        data = _api_get(url, access_token)
        print('drive file ok')
        return data
    except Exception as e:
        print('drive spaces failed: ' + str(e))
    return {}


def dingtalk_read_sheet(access_token):
    """Read spreadsheet data directly via Sheet API."""
    oid = DINGTALK_OPERATOR_ID
    if not oid:
        print('no operatorId, skip sheet read')
        return []
    workbook_id = DINGTALK_NODE_ID
    # Step 1: Get all sheets
    try:
        url = ('https://api.dingtalk.com/v1.0/doc/workbooks/'
               + workbook_id + '/sheets?operatorId=' + oid)
        data = _api_get(url, access_token)
        sheets = data.get('value', data.get('sheets', []))
        if not sheets:
            print('no sheets found in workbook')
            return []
        sheet_id = sheets[0].get('id', sheets[0].get('sheetId', ''))
        print('sheet found: ' + str(sheet_id) + ' (' + str(len(sheets)) + ' sheets)')
    except Exception as e:
        print('get sheets failed: ' + str(e))
        return []
    # Step 2: Read data in chunks (sheet has 5000+ rows, 22 columns)
    all_rows = []
    chunk_size = 1000
    for start in range(1, 10000, chunk_size):
        end = start + chunk_size - 1
        try:
            url = ('https://api.dingtalk.com/v1.0/doc/workbooks/'
                   + workbook_id + '/sheets/' + str(sheet_id)
                   + '/ranges/A' + str(start) + ':V' + str(end)
                   + '?operatorId=' + oid)
            data = _api_get(url, access_token, timeout=30)
            values = data.get('values', [])
            if not values:
                break
            all_rows.extend(values)
            print('read rows ' + str(start) + '-' + str(start + len(values) - 1))
            if len(values) < chunk_size:
                break
        except Exception as e:
            print('read chunk ' + str(start) + ' failed: ' + str(e))
            break
    if not all_rows:
        print('no data in sheet')
        return []
    # Convert to dict rows
    headers = [str(v).strip() if v else '' for v in all_rows[0]]
    print('headers: ' + ', '.join(h for h in headers if h)[:200])
    result = []
    for row_vals in all_rows[1:]:
        if not any(v is not None for v in row_vals):
            continue
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(row_vals):
                val = row_vals[i]
                d[h] = str(val).strip() if val is not None else ''
        if d:
            result.append(d)
    print('sheet rows: ' + str(len(result)))
    return result


def dingtalk_export_download(access_token):
    """Export spreadsheet as xlsx and download."""
    oid = DINGTALK_OPERATOR_ID
    if not oid:
        print('no operatorId, skip export')
        return None
    workbook_id = DINGTALK_NODE_ID
    # Step 1: Submit export job
    try:
        url = ('https://api.dingtalk.com/v1.0/doc/workbooks/'
               + workbook_id + '/export')
        body = {'operatorId': oid, 'targetFormat': 'xlsx'}
        data = _api_post(url, access_token, body)
        task_id = data.get('taskId', '')
        if not task_id:
            print('no taskId in export response')
            return None
        print('export task: ' + task_id)
    except Exception as e:
        print('export submit failed: ' + str(e))
        return None
    # Step 2: Poll for completion
    import time
    file_url = None
    for i in range(30):
        try:
            time.sleep(2)
            url = ('https://api.dingtalk.com/v1.0/doc/workbooks/'
                   + workbook_id + '/export/'
                   + task_id + '?operatorId=' + oid)
            data = _api_get(url, access_token)
            status = data.get('status', '')
            if status in ('completed', 'success', 'done'):
                file_url = data.get('fileUrl', data.get('url', ''))
                break
            elif status in ('failed', 'error'):
                print('export failed: ' + status)
                return None
            print('export status: ' + status + ' (waiting...)')
        except Exception as e:
            print('poll error: ' + str(e))
            break
    if not file_url:
        print('export timeout or no file url')
        return None
    # Step 3: Download file
    try:
        req = urllib.request.Request(file_url)
        with urllib.request.urlopen(req, timeout=60, context=SSL_CTX) as resp:
            file_data = resp.read()
            os.makedirs(DATA_DIR, exist_ok=True)
            fp = os.path.join(DATA_DIR, 'dingtalk_data.xlsx')
            with open(fp, 'wb') as f:
                f.write(file_data)
            print('downloaded: ' + fp + ' (' + str(len(file_data)) + ' bytes)')
            return fp
    except Exception as e:
        print('download failed: ' + str(e))
        return None


def fetch_dingtalk_data():
    if not DINGTALK_APP_KEY or not DINGTALK_APP_SECRET:
        print('no dingtalk credentials, skip')
        return []
    print('--- fetch via DingTalk API ---')
    if not DINGTALK_OPERATOR_ID:
        print('WARNING: DINGTALK_OPERATOR_ID not set')
        print('Please add it to GitHub Secrets (your DingTalk unionId)')
        print('Get it: DingTalk Admin > Contacts > your profile > unionId')
        return []
    try:
        token = dingtalk_get_token()
        # Try reading sheet directly first
        rows = dingtalk_read_sheet(token)
        if rows:
            return rows
        # Fallback: export as xlsx and download
        print('--- trying export/download ---')
        fp = dingtalk_export_download(token)
        if fp:
            return read_xlsx_rows(fp)
    except Exception as e:
        print('dingtalk API error: ' + str(e))
    return []


def read_xlsx_rows(path):
    if not openpyxl:
        print('openpyxl not installed')
        return []
    rows = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for c in ws[1]:
            headers.append(str(c.value).strip() if c.value else '')
        col_idx = {}
        for i, h in enumerate(headers):
            if h:
                col_idx[h] = i
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            d = {}
            for name, idx in col_idx.items():
                val = row[idx] if idx < len(row) else None
                d[name] = str(val).strip() if val is not None else ''
            rows.append(d)
        wb.close()
        print(os.path.basename(path) + ': ' + str(len(rows)) + ' rows')
    except Exception as e:
        print('read error ' + os.path.basename(path) + ': ' + str(e))
    return rows


def read_local_data():
    all_rows = []
    if not os.path.exists(DATA_DIR):
        return all_rows
    for fn in sorted(os.listdir(DATA_DIR)):
        fp = os.path.join(DATA_DIR, fn)
        if fn.endswith('.csv'):
            try:
                with open(fp, encoding='utf-8-sig') as f:
                    first = f.readline()
                    if not first or '\x00' in first:
                        continue
                    f.seek(0)
                    for row in csv.DictReader(f):
                        d = {}
                        for k, v in row.items():
                            d[k.strip() if k else ''] = (v.strip() if v else '')
                        all_rows.append(d)
                print(fn + ': ' + str(len(all_rows)) + ' rows')
            except Exception as e:
                print(fn + ' error: ' + str(e))
        elif fn.endswith('.xlsx') and openpyxl:
            all_rows.extend(read_xlsx_rows(fp))
    return all_rows


def parse_date(s):
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S',
                '%Y/%m/%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    try:
        serial = float(s)
        dt = datetime(1899, 12, 30) + timedelta(days=serial)
        return dt.strftime('%Y-%m-%d')
    except (ValueError, OverflowError):
        pass
    return None


def parse_float(s, default=0.0):
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def is_expired(b):
    if b['status'] != CLOSED:
        return False
    if b.get('over90') != OVER90:
        return False
    cd = parse_date(b.get('created'))
    if cd:
        days = (datetime.now() - datetime.strptime(cd, '%Y-%m-%d')).days
        if days > 90:
            return True
    return False


def process_rows(rows):
    bugs = []
    for r in rows:
        # Project: try DB-项目 (actual column name)
        project = ''
        for k in ('DB-\u9879\u76ee', 'DI-\u9879\u76ee'):
            v = r.get(k, '').strip()
            if v:
                project = v
                break
        if not project:
            project = GENERAL
        di_val = parse_float(r.get('DB-DI\u503c', '0'))
        status = r.get('DB-\u4efb\u52a1\u72b6\u6001', '').strip()
        created = parse_date(r.get('\u521b\u5efa\u65f6\u95f4', ''))
        di_date = parse_date(r.get('DB-Date', '') or r.get('DI-Date', ''))
        sla_val = r.get('DB-SLA\u8d85\u65f6', '').strip()
        bugs.append({
            'id': r.get('\u4efb\u52a1ID', '').strip() or r.get('\u7f3a\u9677ID', '').strip(),
            'title': r.get('\u6807\u9898', '').strip(),
            'level': r.get('DB-BUG\u7b49\u7ea7', '').strip() or r.get('BUG\u7b49\u7ea7', '').strip() or 'none',
            'status': status,
            'assignee': r.get('\u6267\u884c\u8005', '').strip(),
            'resolver': r.get('\u89e3\u51b3\u8005', '').strip(),
            'created': created or '',
            'di_date': di_date or '',
            'di_value': di_val,
            'project': project,
            'owner': r.get('DB-Owner', '').strip(),
            'dept': r.get('DB-\u90e8\u95e8', '').strip(),
            'role': r.get('DB-\u89d2\u8272', '').strip(),
            'sla_timeout': 1 if sla_val else 0,
            'sla_days': parse_float(r.get('DB-SLA\u5929\u6570', '0')),
            'over90': r.get('\u89e3\u51b3\u8d8590\u5929', '').strip(),
            'risk_ctrl': r.get('\u91ca\u653e\u7ba1\u63a7', '').strip(),
        })
    return bugs


def compute_projects_data(bugs):
    by_project = defaultdict(list)
    for b in bugs:
        if is_expired(b):
            continue
        by_project[b['project']].append(b)
    projects_data = {}
    for proj in sorted(by_project.keys()):
        pb = by_project[proj]
        open_di = round(sum(b['di_value'] for b in pb
                            if b['status'] in OPEN_SET), 2)
        resolved_cnt = sum(1 for b in pb if b['status'] in CLOSED_SET)
        total = len(pb)
        sla = sum(1 for b in pb if b['sla_timeout'] > 0)
        block = sum(1 for b in pb if b['status'] not in CLOSED_SET
                    and b['risk_ctrl'] == HIGH_RISK)
        rr = '{:.1f}'.format(resolved_cnt / total * 100) if total > 0 else '0.0'
        preg = sum(1 for b in pb if b['status'] == REOPEN)
        if open_di < 20 and sla <= 2:
            health = 'normal'
        elif open_di < 40 and sla <= 5:
            health = 'warning'
        elif open_di < 80 and sla <= 20:
            health = 'caution'
        else:
            health = 'danger'
        js_bugs = []
        for b in pb:
            bod = b['di_value'] if b['status'] in OPEN_SET else 0.0
            js_bugs.append({
                'id': b['id'], 'title': b['title'], 'level': b['level'],
                'status': b['status'], 'assignee': b['assignee'],
                'created': b['created'], 'openDI': round(bod, 2),
                'sla_timeout': b['sla_timeout'], 'sla_days': b['sla_days'],
                'db_dept': b['dept'], 'db_role': b['role'],
            })
        projects_data[proj] = {
            'bugs': js_bugs,
            'stats': {
                'health': health, 'open_di': open_di,
                'resolve_rate': rr, 'pending_regression': preg,
                'sla': sla, 'block': block,
                'resolved': resolved_cnt, 'total': total,
            }
        }
    return projects_data


def compute_trends(bugs, projects_data):
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    dates = [(today - timedelta(days=i)).strftime('%Y-%m-%d')
             for i in range(TREND_DAYS - 1, -1, -1)]
    labels = [d[5:] for d in dates]
    valid = [b for b in bugs if not is_expired(b)]
    dnew = []
    dres = []
    for d in dates:
        dnew.append(round(sum(b['di_value'] for b in valid
                             if b['created'] == d), 2))
        dres.append(round(sum(b['di_value'] for b in valid
                              if b['status'] in CLOSED_SET
                              and b['di_date'] == d), 2))
    cum = []
    for i, d in enumerate(dates):
        if i == 0:
            cum.append(round(sum(b['di_value'] for b in valid
                                 if b['created'] <= d
                                 and b['status'] in OPEN_SET), 2))
        else:
            cum.append(round(cum[i - 1] + dnew[i] - dres[i], 2))
    pt = {}
    for proj in sorted(projects_data.keys()):
        pb = [b for b in valid if b['project'] == proj]
        pn = []
        pr = []
        for d in dates:
            pn.append(round(sum(b['di_value'] for b in pb
                                if b['created'] == d), 2))
            pr.append(round(sum(b['di_value'] for b in pb
                                if b['status'] in CLOSED_SET
                                and b['di_date'] == d), 2))
        pc = []
        for i, d in enumerate(dates):
            if i == 0:
                pc.append(round(sum(b['di_value'] for b in pb
                                    if b['created'] <= d
                                    and b['status'] in OPEN_SET), 2))
            else:
                pc.append(round(pc[i - 1] + pn[i] - pr[i], 2))
        pt[proj] = [pc, pn, pr]
    return {'dates': labels, 'cumulative': cum, 'daily_new': dnew,
            'daily_resolved': dres, 'project_trends': pt}


def compute_person_stats(bugs, projects_data):
    valid = [b for b in bugs if not is_expired(b)]
    pm = {}
    for b in valid:
        owner = b['owner'] or b['assignee']
        if not owner:
            continue
        if owner not in pm:
            pm[owner] = {
                'name': owner, 'role': b['role'] or UNKNOWN,
                'proj_contrib': {},
                'total_open_di': 0, 'total_solved_di': 0,
                'total_unsolved_di': 0, 'total_sla': 0,
                'total_reopen': 0, 'solved_tickets': 0,
                'total_tickets': 0, 'sla_days_sum': 0,
                'sla_days_count': 0,
            }
        p = pm[owner]
        p['total_tickets'] += 1
        if b['status'] in CLOSED_SET:
            p['solved_tickets'] += 1
            p['total_solved_di'] += b['di_value']
        elif b['status'] == REOPEN:
            p['total_reopen'] += 1
            p['total_unsolved_di'] += b['di_value']
        else:
            p['total_unsolved_di'] += b['di_value']
            if b['status'] in OPEN_SET:
                p['total_open_di'] += b['di_value']
        p['total_sla'] += b['sla_timeout']
        if b['sla_days'] > 0:
            p['sla_days_sum'] += b['sla_days']
            p['sla_days_count'] += 1
        proj = b['project']
        if proj not in p['proj_contrib']:
            p['proj_contrib'][proj] = {'solved_di': 0, 'open_di': 0,
                                       'unsolved_di': 0}
        if b['status'] in CLOSED_SET:
            p['proj_contrib'][proj]['solved_di'] += b['di_value']
        elif b['status'] == REOPEN:
            p['proj_contrib'][proj]['unsolved_di'] += b['di_value']
        else:
            p['proj_contrib'][proj]['unsolved_di'] += b['di_value']
            if b['status'] in OPEN_SET:
                p['proj_contrib'][proj]['open_di'] += b['di_value']
    for p in pm.values():
        p['total_open_di'] = round(p['total_open_di'], 2)
        p['total_solved_di'] = round(p['total_solved_di'], 2)
        p['total_unsolved_di'] = round(p['total_unsolved_di'], 2)
        for proj in p['proj_contrib']:
            for k in p['proj_contrib'][proj]:
                p['proj_contrib'][proj][k] = round(
                    p['proj_contrib'][proj][k], 2)
    return list(pm.values())


def read_milestones():
    path = os.path.join('milestone', '\u91cc\u7a0b\u7891\u8282\u70b9.xlsx')
    if not os.path.exists(path) or not openpyxl:
        return {}
    ms = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            items = []
            for row in rows[1:]:
                if not any(v is not None for v in row):
                    continue
                dv = row[0] if len(row) > 0 else None
                nv = row[1] if len(row) > 1 else None
                sv = row[2] if len(row) > 2 else ''
                rv = row[3] if len(row) > 3 else ''
                if nv is None:
                    continue
                ds = ''
                if dv:
                    try:
                        serial = float(dv)
                        ds = (datetime(1899, 12, 30)
                              + timedelta(days=serial)).strftime('%Y-%m-%d')
                    except (ValueError, TypeError):
                        ds = str(dv)
                items.append({
                    'date': ds, 'name': str(nv).strip(),
                    'item': str(rv).strip() if rv else '',
                    'status': str(sv).strip() if sv else IN_PROGRESS,
                })
            ms[sn] = items
        wb.close()
        print('milestones: ' + str(sum(len(v) for v in ms.values())))
    except Exception as e:
        print('milestone error: ' + str(e))
    return ms


def read_version_plan():
    path = os.path.join('Version_Plan', 'Version_Plan.xlsx')
    if not os.path.exists(path) or not openpyxl:
        return []
    plans = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            for row in rows[1:]:
                if not any(v is not None for v in row):
                    continue
                plans.append({
                    'date': str(row[0]).strip() if row[0] else '',
                    'name': sn,
                    'overview': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                    'link': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                })
        wb.close()
        print('version plans: ' + str(len(plans)))
    except Exception as e:
        print('version plan error: ' + str(e))
    return plans


def inject_data(html_path, data):
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    reps = {
        'TREND_DATES': data['trend_dates'],
        'CUMULATIVE_DI': data['cumulative_di'],
        'DAILY_NEW': data['daily_new'],
        'DAILY_RESOLVED': data['daily_resolved'],
        'PROJECTS_DATA': data['projects_data'],
        'PROJECT_NAMES': data['project_names'],
        'PROJECT_TRENDS': data['project_trends'],
        'MILESTONES': data['milestones'],
        'PERSON_STATS': data['person_stats'],
        'VERSION_PLAN': data['version_plan'],
    }
    for var_name, value in reps.items():
        js = 'const ' + var_name + ' = ' + json.dumps(value, ensure_ascii=False) + ';'
        pat = r'const\s+' + var_name + r'\s*=\s*[\s\S]*?;'
        if re.search(pat, html):
            html = re.sub(pat, js, html)
        else:
            html = html.replace('<script>\n', '<script>\n' + js + '\n', 1)
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)


def main():
    print('=' * 50)
    print('TPM Dashboard Build')
    print('=' * 50)
    rows = fetch_dingtalk_data()
    if not rows:
        print('error: no data from DingTalk API')
        return
    print('total rows: ' + str(len(rows)))
    bugs = process_rows(rows)
    print('valid bugs: ' + str(len(bugs)))
    projects_data = compute_projects_data(bugs)
    print('projects: ' + str(len(projects_data)))
    trends = compute_trends(bugs, projects_data)
    person_stats = compute_person_stats(bugs, projects_data)
    milestones = read_milestones()
    version_plan = read_version_plan()
    project_names = sorted(projects_data.keys())
    for proj in project_names:
        s = projects_data[proj]['stats']
        print(proj + ': OPEN DI=' + str(s['open_di'])
              + ', SLA=' + str(s['sla'])
              + ', BLOCK=' + str(s['block'])
              + ', rate=' + str(s['resolve_rate'])
              + '%, health=' + s['health'])
    print('staff: ' + str(len(person_stats)))
    print('trend days: ' + str(len(trends['dates'])))
    html_path = 'index.html'
    if not os.path.exists(html_path):
        print('error: ' + html_path + ' not found')
        return
    inject_data(html_path, {
        'trend_dates': trends['dates'],
        'cumulative_di': trends['cumulative'],
        'daily_new': trends['daily_new'],
        'daily_resolved': trends['daily_resolved'],
        'projects_data': projects_data,
        'project_names': project_names,
        'project_trends': trends['project_trends'],
        'milestones': milestones,
        'person_stats': person_stats,
        'version_plan': version_plan,
    })
    print('index.html generated successfully')
    print('path: ' + os.path.abspath(html_path))


if __name__ == '__main__':
    main()
