#!/usr/bin/env python3
"""
数据注入脚本：自动扫描 data 目录下的所有 CSV/XLSX 文件并生成看板 HTML
"""
import csv, json, re, os
import openpyxl

DATA_DIR = 'data'

# DI 权重映射
DI_WEIGHTS = {'P0': 10.0, 'P1': 3.0, 'P2': 1.0, 'P3': 0.1}

def auto_discover_projects():
    """自动扫描 data 目录下的所有 CSV/XLSX 文件"""
    projects = []
    if not os.path.exists(DATA_DIR):
        return projects
    
    for filename in sorted(os.listdir(DATA_DIR)):
        if filename.endswith('.csv') or filename.endswith('.xlsx'):
            proj_name = filename.replace('.csv', '').replace('.xlsx', '')
            projects.append(proj_name)
    
    return projects

def calc_open_di(level, status):
    """计算 OPEN DI 值：已关闭/已解决不计分，已解决乘以50%"""
    base = DI_WEIGHTS.get(level, 0)
    if status in ('已关闭', '已解决'):
        return 0.0
    return base

def read_csv(path):
    """读取单个 CSV 文件"""
    bugs = []
    try:
        with open(path, encoding='utf-8-sig') as f:
            # 检查是否为有效文本文件
            first_line = f.readline()
            if not first_line or '\x00' in first_line:
                print(f'{os.path.basename(path)}: 非文本文件，跳过')
                return bugs
            f.seek(0)
            
            reader = csv.DictReader(f)
            for row in reader:
                level = row.get('BUG等级','').strip() or 'none'
                status = row.get('任务状态','').strip()
                open_di = float(row.get('OPEN DI值','0') or 0)
                total_di = float(row.get('DI总值','0') or 0)
                
                # 如果 CSV 中没有 DI 值列，则自动计算
                if open_di == 0 and total_di == 0:
                    total_di = DI_WEIGHTS.get(level, 0)
                    open_di = calc_open_di(level, status)
                
                bugs.append({
                    'id': row.get('任务ID','').strip(),
                    'title': row.get('标题','').strip(),
                    'level': level,
                    'status': status,
                    'assignee': row.get('执行者','').strip(),
                    'createdDate': row.get('创建时间','').strip(),
                    'openDI': open_di,
                    'totalDI': total_di,
                })
        print(f'{os.path.basename(path)}: {len(bugs)} bugs')
    except Exception as e:
        print(f'{os.path.basename(path)} error: {e}')
    return bugs

def read_xlsx(path):
    """读取单个 XLSX 文件"""
    bugs = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        
        # 构建表头索引
        headers = [cell.value for cell in ws[1]]
        col_idx = {}
        for i, h in enumerate(headers):
            if h:
                col_idx[str(h).strip()] = i
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            
            def get_col(name, default=''):
                idx = col_idx.get(name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else default
                return default
            
            level = get_col('BUG等级', 'none')
            status = get_col('任务状态', '')
            assignee = get_col('执行者', '')
            
            # XLSX 没有 OPEN DI / DI 总值列，自动计算
            total_di = DI_WEIGHTS.get(level, 0)
            open_di = calc_open_di(level, status)
            
            bugs.append({
                'id': get_col('任务ID'),
                'title': get_col('标题'),
                'level': level,
                'status': status,
                'assignee': assignee,
                'createdDate': get_col('创建时间'),
                'openDI': open_di,
                'totalDI': total_di,
            })
        
        wb.close()
        print(f'{os.path.basename(path)}: {len(bugs)} bugs')
    except Exception as e:
        print(f'{os.path.basename(path)} error: {e}')
    return bugs

def excel_date_to_str(excel_serial):
    """将Excel日期序列号转换为 YYYY-MM-DD 字符串"""
    try:
        serial = float(excel_serial)
        # Excel 1900日期系统：1 = 1900-01-01
        from datetime import datetime, timedelta
        base = datetime(1899, 12, 30)  # Excel的基准日
        dt = base + timedelta(days=serial)
        return dt.strftime('%Y-%m-%d')
    except:
        return str(excel_serial)

def read_milestone_data():
    """读取里程碑数据"""
    milestone_path = os.path.join(os.path.dirname(__file__), 'milestone', '里程碑节点.xlsx')
    milestones = {}
    
    if not os.path.exists(milestone_path):
        print('里程碑文件不存在，跳过')
        return milestones
    
    try:
        wb = openpyxl.load_workbook(milestone_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            
            items = []
            for row in rows[1:]:  # 跳过表头
                if not any(v is not None for v in row):
                    continue
                date_val = row[0] if len(row) > 0 else None
                name_val = row[1] if len(row) > 1 else None
                status_val = row[2] if len(row) > 2 else ''
                risk_val = row[3] if len(row) > 3 else ''
                
                if name_val is None:
                    continue
                
                date_str = excel_date_to_str(date_val) if date_val else ''
                items.append({
                    'date': date_str,
                    'name': str(name_val).strip(),
                    'status': str(status_val).strip() if status_val else '',
                    'risk': str(risk_val).strip() if risk_val else ''
                })
            
            milestones[sheet_name] = items
            print(f'里程碑 {sheet_name}: {len(items)} 个节点')
        
        wb.close()
    except Exception as e:
        print(f'里程碑读取错误: {e}')
    
    return milestones

def extract_project_risks(milestones):
    """从里程碑数据中提取每个项目的风险项列表"""
    risks = {}
    for proj_id, items in milestones.items():
        proj_risks = [item['risk'] for item in items if item.get('risk')]
        risks[proj_id] = proj_risks
    return risks

def main():
    all_projects = []
    
    # 自动发现项目
    projects = auto_discover_projects()
    if not projects:
        print('错误：data 目录下没有找到任何 CSV/XLSX 文件')
        return
    
    # 遍历项目，读取对应的文件
    for proj in projects:
        csv_path = os.path.join(DATA_DIR, f'{proj}.csv')
        xlsx_path = os.path.join(DATA_DIR, f'{proj}.xlsx')
        
        bugs = []
        if os.path.exists(csv_path):
            bugs = read_csv(csv_path)
        elif os.path.exists(xlsx_path):
            bugs = read_xlsx(xlsx_path)
        else:
            print(f'{proj} 对应文件不存在，跳过')
            continue
        
        if bugs:
            all_projects.append({'id': proj, 'name': proj, 'bugs': bugs})
    
    if not all_projects:
        print('错误：没有找到任何有效的数据文件')
        return
    
    # 生成 JS 数据
    js_data = 'const PROJECTS = ' + json.dumps(all_projects, ensure_ascii=False) + ';'
    
    # 读取里程碑数据
    milestones = read_milestone_data()
    milestone_js = 'const MILESTONES = ' + json.dumps(milestones, ensure_ascii=False) + ';'
    
    # 提取项目风险项
    project_risks = extract_project_risks(milestones)
    risks_js = 'const PROJECT_RISKS = ' + json.dumps(project_risks, ensure_ascii=False) + ';'
    
    # 读取 HTML 模板
    html_path = 'index.html'
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    
    # 替换 PROJECTS 数据
    new_html = re.sub(r'const PROJECTS = \[.*?\];', js_data, html, flags=re.DOTALL)
    
    # 检查是否已有 MILESTONES 和 PROJECT_RISKS 变量，有则替换，无则在 PROJECTS 后追加
    if re.search(r'const MILESTONES = ', new_html):
        new_html = re.sub(r'const MILESTONES = \{.*?\};', milestone_js, new_html, flags=re.DOTALL)
    else:
        # 在 PROJECTS 变量后面追加
        new_html = re.sub(r'(const PROJECTS = \[.*?\];)', r'\1\n' + milestone_js, new_html, flags=re.DOTALL)
    
    if re.search(r'const PROJECT_RISKS = ', new_html):
        new_html = re.sub(r'const PROJECT_RISKS = \{.*?\};', risks_js, new_html, flags=re.DOTALL)
    else:
        # 在 MILESTONES 变量后面追加
        new_html = re.sub(r'(const MILESTONES = \{.*?\};)', r'\1\n' + risks_js, new_html, flags=re.DOTALL)
    
    # 写回
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(new_html)
    
    total_bugs = sum(len(p['bugs']) for p in all_projects)
    total_milestones = sum(len(v) for v in milestones.values())
    print(f'\n✓ 数据注入完成')
    print(f'  - Bug记录: {total_bugs} 条')
    print(f'  - 里程碑节点: {total_milestones} 个 ({len(milestones)} 个项目)')
    print(f'  - 看板路径: {os.path.abspath(html_path)}')

if __name__ == '__main__':
    main()
